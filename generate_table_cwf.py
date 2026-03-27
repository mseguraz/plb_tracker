import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

#Merge two list without duplicates
def merge_lists_no_duplicates(list1, list2):
    merged_set = set(list1 + list2)
    return sorted(merged_set)

#Create plist tracking table
def generate_excel_tables(data: dict, output_file: str, names: list):
    wb = openpyxl.load_workbook('/nfs/site/disks/mdo_dcd_001/scan/jspereir/Scripts/gnr_plb_tracker/reference.xlsx')

    # Select the sheet where you want to append data (adjust the sheet name accordingly)
    sheet = wb["Sheet1"]  # Change "Sheet1" to your actual sheet name
    current_columns=None

    # Create a workbook and add sheets
    #wb = openpyxl.Workbook()
    #wb.remove(wb.active)  # Remove the default sheet
    
    # Define colors
    green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid") #In kill
    yellow_fill = PatternFill(start_color="FFFF50", end_color="FFFF50", fill_type="solid") #In EDC
    gold_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid") #Not in flow
    
    #get column names
    for _, columns in data.items():
        col_names = list(columns.keys())

        if not current_columns:
            current_columns=col_names 
        current_columns=merge_lists_no_duplicates(current_columns, col_names)
        #print(prev_columns, '\n', table_name, current_columns, sheet.max_row)

    i=0
    sheet.title=output_file
    sheet.append([""])
    for table_name, columns in data.items():
        #print(table_name, names[i])
        col_names=current_columns
        #print(col_names)
        sheet.append(["", ""] + col_names)  # Create header row
        
        # Generate the 'version' row
        version_row = ['EDT']
        for col_name in col_names:
            if col_name in columns:
                first_row_key = next(iter(columns[col_name]))
                version_row.append(columns[col_name][first_row_key][0])
            else:
                version_row.append(None)

        #print(version_row)
        sheet.append([""]+version_row)

        # Get the row with EDT (rev) data
        edt_row = sheet.max_row
        #print(edt_row)

        # Determine unique rows excluding 'version'
        all_rows = sorted(set(row for col in columns.values() for row in col))
        #print(all_rows)
        for row_name in all_rows:
            row_data = [row_name]
            for col_name in col_names:
                if col_name in columns:
                    if row_name in columns[col_name]:
                        str_value, int_value, condition, mask, _ = columns[col_name][row_name]
                        #print([col_name, row_name], str_value, int_value, condition)
                        
                        # Determine cell value based on the 'version' row
                        version_value = sheet.cell(row=edt_row, column=col_names.index(col_name) + 3).value
                        #print('here', version_value)
                        if str_value != version_value:
                            cell_value = f"{int_value}_{str_value}" if not mask else f"*{int_value}_{str_value}"
                        else:
                            cell_value = int_value if not mask else f"*{int_value}"
                        
                        # Add cell value to row data
                        row_data.append(cell_value)
                    else:
                        row_data.append(None)  # Empty cell for missing row
                else:
                    row_data.append(None)  # Empty cell for missing row
            #print('here2', row_data)
            sheet.append([""]+row_data)
            for cell in sheet[sheet.max_row]:
                cell.alignment = Alignment(horizontal='right')
        
        # Apply colors
        #print(table_name, edt_row, sheet.max_row, sheet.max_column)
        for row in sheet.iter_rows(min_row=edt_row+1, max_row=sheet.max_row, min_col=3, max_col=sheet.max_column):
            for cell in row:
                #print('row: ', row, 'cell: ', cell)
                if cell.value is None:
                    cell.fill = gold_fill
                else:
                    col_name = sheet.cell(row=edt_row-1, column=cell.column).value
                    row_name = sheet.cell(row=cell.row, column=2).value
                    #print('here', col_name, row_name)
                    if col_name in columns and row_name in columns[col_name]:
                        condition = columns[col_name][row_name][2]
                        cell.fill = yellow_fill if condition else green_fill
                    else:
                        cell.fill = gold_fill  # Just in case, though this should be covered by the earlier check
    
        # Apply bold to the entire column B
        for cell in sheet['B']:
            cell.font = Font(bold=True)
                     
        # Get the last row
        last_row = sheet.max_row
        sheet.merge_cells(f'A{edt_row}:A{last_row}')
        sheet[f'A{edt_row}']=names[i].upper()
        sheet[f'A{edt_row}'].alignment = Alignment(vertical='center', horizontal='center', textRotation=90)
        sheet[f'A{edt_row}'].font = Font(bold=True, size=18)

        # Apply bold to header row
        for cell in sheet[edt_row-1]:
            #cell.alignment = Alignment(vertical='center', horizontal='center', textRotation=90)
            cell.font =  Font(bold=True)

        i+=1
        sheet.append([""])
        
    wb.save(output_file+'.xlsx')

#Plist tracking table for audit
def generate_excel_tables_audit(data: dict, output_file: str, bin_dict: dict):
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    first_sheet = True
   
    
    # Define colors
    blue_fill = PatternFill(start_color="4D93D9", end_color="4D93D9", fill_type="solid") 
    wrap_alignment = Alignment(wrapText=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"))
    
    #get column names
    col_names = ['INSTANCE', 'BIN', 'PLIST', 'PARTITIONS', 'GROUPING', 'PHASES', 'STEPPING', 'REVISION', 'CHUNKS']

    for names in data:
         # Create a new sheet with the given name
        if first_sheet:
            ws = wb.active  # Use the first sheet (default one created)
            ws.title = names
            first_sheet = False
        else:
            ws = wb.create_sheet(title=names)
        
       
        ws.append([""])
        ws.append([""] + col_names)  # Create header row

        # Apply bold to header row
        for col in range(2, len(col_names) + 2):
            cell = ws.cell(row=2, column=col)
            cell.font =  Font(bold=True)
            cell.fill = blue_fill
            cell.border = thin_border

        start_row = 2
        og=3
        max_length = 0
        this=0
    
        if not data[names]:
            continue
        for instance, columns in data[names].items():
            #print(instance, columns)
            try:
                softbin=bin_dict[instance]
                col_12=[softbin]
                #print(instance, softbin)
            except:
                col_12=[""]
            
            col1=[instance]
            col2=[columns[0]]

            text=""
            for p in columns[1]:
                text+=p
                text+='\n'

            col3=[text[:-1]]
            
            try:
                this=max([len(thing) for thing in columns[1]]) 
            except:
                this=len(columns[0]) if len(columns[0])>this else this
                
            max_length=this if this>max_length else max_length
            values=columns[2]
            #print(values)
            if not values:
                print(F"Instance {instance} has no values")
                continue

            col4=[]
            col5=[]
            col6=[]
            col7=[]
            extra_col=[]

            for key in list(values.keys()):
                col4=col4+[key]+[""]*(len(list(values[key].keys()))-1)
                col5+=list(values[key].keys())
        
            lenghts=[]
            for key1 in list(values.keys()):
                lenghts.append(len(list(values[key1].keys())))
                for key2 in list(values[key1].keys()):
                    col6+=[values[key1][key2][0]]
                    col7+=[values[key1][key2][1]]
                    extra_col+=[values[key1][key2][4]]

            #print(lenghts)
            len_instance=max(len(col4), len(col5), len(col6), len(col7))

            col1+=[""]*(len_instance-len(col1))
            col_12+=[""]*(len_instance-len(col_12))
            col2+=[""]*(len_instance-len(col2))
            col3+=[""]*(len_instance-len(col3))
            col4+=[""]*(len_instance-len(col4))
            col5+=[""]*(len_instance-len(col5))
            col6+=[""]*(len_instance-len(col6))
            col7+=[""]*(len_instance-len(col7))
            extra_col+=[""]*(len_instance-len(extra_col))

            #print(col1)
            #print(col2)
            #print(col3)
            #print(col4)
            #print(col5)
            #print(col6)
            #print(col7)

            for i in range(len_instance):
                row=["", col1[i], col_12[i], col2[i], col3[i], col4[i], col5[i], extra_col[i], col6[i], col7[i]]
                ws.append(row)
                current_row = ws.max_row
                ws.cell(row=current_row, column=5).alignment = wrap_alignment

                for col in range(1, len(row) + 1):  # Offset by 1
                    ws.cell(row=current_row, column=col).border = thin_border  # Apply border

            # Merge column A for the length of the table and add "Table" vertically
            end_row = ws.max_row
            #merge cells
            ws.merge_cells(start_row=start_row+1, start_column=2, end_row=end_row, end_column=2) #for instance
            ws.merge_cells(start_row=start_row+1, start_column=3, end_row=end_row, end_column=3) #for plist
            ws.merge_cells(start_row=start_row+1, start_column=4, end_row=end_row, end_column=4) #for partitions
            ws.merge_cells(start_row=start_row+1, start_column=5, end_row=end_row, end_column=5) #for partitions

            merged_cell = ws.cell(row=start_row+1, column=1)
            merged_cell.value = names
            merged_cell.font = Font(bold=True, size=16)
            merged_cell.alignment = Alignment(horizontal="left", vertical="center", text_rotation=90)  # Rotate text vertically

            #for grouping
            temp_pos=start_row+1
            for l in lenghts:
                ws.merge_cells(start_row=temp_pos, start_column=6, end_row=temp_pos+l-1, end_column=6)
                temp_pos+=l

            start_row=ws.max_row
            
        end_row = ws.max_row
        #print(og, end_row, instance, values)
        try:
            ws.merge_cells(start_row=og, start_column=1, end_row=end_row, end_column=1) #for flow
        except:
            continue
        og=ws.max_row+1

        ws.column_dimensions['B'].width = max_length  # Set max width (50)
        ws.column_dimensions['D'].width = max_length  # Set max width (50)
        ws.column_dimensions['E'].width = max_length  # Set max width (50)
        #print(max_length)
    wb.save(output_file)
        
# Example usage
data = {
    'begin': {
        'tile1': {'a': ['r0', 1, False], 'b': ['r0', 2, True], 'c': ['r0', 3, False]},
        'tile2': {'a': ['r0', 2, False], 'b': ['r1', 3, True], 'c': ['r1', 3, False]}, 
    },
    'end': {
        'tile1': {'w': ['r1', 1, True]},
        'tile2': {'c': ['r0', 3, False], 'd': ['r0', 3, False]},
    },
}

data2 = {
    'ATPG_CHATOP_HRY_K_BEGIN_STF_CFC_NOM_LFM_IE_NR_EDC': 
        [
            'scn_s_cfc_x_begin_sEs_edt_cputopnrllc_atpg_TTR_list', 
            ['casa', 'paloto'], 
            {'cputopnr': {'ph1': ['r2', 17, 'T', False], 'ph2': ['r2', 4, 'T', False], 'ph3': ['r2', 1, 'T', False]},
             'cpubot':{'ph1': ['r1', 16, 'T', False], 'ph2': ['r2', 4, 'T', False], 'ph3': ['r0', 1, 'T', False]}}],
    'ATPG_CHATOP_HRY_K_BEGIN_STF_CFC_NOM_LFM_IE': 
        [
            'scn_s_cfc_x_begin_sEs_edt_cputopall_atpg_TTR_list', 
            [], 
            {'cputopall': {'ph1': ['r0', 17, 'T', False], 'ph2': ['r0', 4, 'T', False], 'ph3': ['r0', 1, 'T', False]}}]
        }

#generate_excel_tables(data, "output_table.xlsx")

#generate_excel_tables_audit(data2, "output_table.xlsx", 'flow')
